"""Verify the imported DB reproduces the source workbook exactly.

For every position we recompute market value and (for stocks) the class /
portfolio / active weights from the DB, then compare against the values stored
in the original sheet. Any cell off by more than the tolerance is reported.

Usage:  python -m scripts.reconcile        (exit code 0 = PASS, 1 = FAIL)
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import numpy as np  # noqa: E402
import openpyxl  # noqa: E402
import pandas as pd  # noqa: E402

from src.config import db_path, load_config, workbook_path  # noqa: E402
from src.model import db as _db  # noqa: E402
from src.model.schema import get_connection  # noqa: E402

MV_TOL = 0.01      # dollars
WEIGHT_TOL = 1e-6  # fraction

# sheet columns (1-based): G=mv, K=class, L=portfolio, M=active
SHEET_COL = {"mv": 7, "class_w": 11, "port_w": 12, "active_w": 13}


def _db_frame(conn, fund: str) -> pd.DataFrame:
    cur = conn.execute(
        _db.q(conn, """
        SELECT h.ticker, s.sec_type, h.shares, h.passive_weight,
               (SELECT close FROM prices p WHERE p.ticker = h.ticker
                AND p.source = 'xlsx_snapshot' ORDER BY date DESC LIMIT 1) AS price
        FROM holdings h JOIN securities s ON s.ticker = h.ticker
        WHERE h.fund = ?
        """),
        (fund,),
    )
    cols = [d[0] for d in cur.description]
    df = pd.DataFrame(cur.fetchall(), columns=cols)
    df["mv"] = df["shares"] * df["price"]

    total = df["mv"].sum()
    cash = df.loc[df.sec_type == "cash", "mv"].sum()
    index = df.loc[df.sec_type == "etf", "mv"].sum()
    invested = total - cash            # excl. cash (incl. index)
    active_sleeve = total - cash - index  # active stock sleeve only

    is_stock = df.sec_type == "stock"
    df["class_w"] = np.where(is_stock, df["mv"] / active_sleeve, np.nan)
    df["port_w"] = np.where(is_stock, df["mv"] / invested, np.nan)
    df["active_w"] = np.where(is_stock, df["port_w"] - df["passive_weight"], np.nan)
    df.attrs["total"] = total
    return df


def _sheet_map(ws) -> dict[str, dict]:
    out: dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        name, ticker = ws.cell(r, 3).value, ws.cell(r, 4).value
        if not name or not ticker:
            continue
        out[str(ticker).strip()] = {
            k: ws.cell(r, c).value for k, c in SHEET_COL.items()
        }
    return out


def reconcile(cfg: dict) -> dict:
    wb = openpyxl.load_workbook(workbook_path(cfg), data_only=True)
    conn = get_connection(db_path(cfg))
    report = {"funds": {}, "ok": True}

    try:
        for fund in cfg["funds"]:
            name = fund["name"]
            db = _db_frame(conn, name)
            sheet = _sheet_map(wb[fund["sheet"]])

            diffs = {"mv": [], "class_w": [], "port_w": [], "active_w": []}
            mismatches = []
            for _, row in db.iterrows():
                tk = row["ticker"]
                src = sheet.get(tk, {})
                for field, tol in (("mv", MV_TOL), ("class_w", WEIGHT_TOL),
                                   ("port_w", WEIGHT_TOL), ("active_w", WEIGHT_TOL)):
                    a, b = row[field], src.get(field)
                    if pd.isna(a) or b is None or not isinstance(b, (int, float)):
                        continue
                    d = abs(float(a) - float(b))
                    diffs[field].append(d)
                    if d > tol:
                        mismatches.append((tk, field, float(a), float(b), d))

            sheet_total = sum(
                v["mv"] for v in sheet.values() if isinstance(v["mv"], (int, float))
            )
            total_diff = abs(db.attrs["total"] - sheet_total)
            ok = not mismatches and total_diff <= MV_TOL
            report["ok"] = report["ok"] and ok
            report["funds"][name] = {
                "n": len(db),
                "db_total": db.attrs["total"],
                "sheet_total": sheet_total,
                "total_diff": total_diff,
                "max": {f: (max(v) if v else 0.0) for f, v in diffs.items()},
                "mismatches": mismatches,
                "ok": ok,
            }
    finally:
        conn.close()
    return report


def main() -> int:
    cfg = load_config()
    report = reconcile(cfg)

    for name, r in report["funds"].items():
        flag = "PASS" if r["ok"] else "FAIL"
        print(f"\n[{flag}] {name}  ({r['n']} positions)")
        print(f"   total MV   db=${r['db_total']:,.2f}  sheet=${r['sheet_total']:,.2f}"
              f"  diff=${r['total_diff']:.4f}")
        m = r["max"]
        print(f"   max |diff| mv=${m['mv']:.4f}  class={m['class_w']:.2e}"
              f"  port={m['port_w']:.2e}  active={m['active_w']:.2e}")
        for tk, field, a, b, d in r["mismatches"]:
            print(f"   ! {tk} {field}: db={a:.6f} sheet={b:.6f} diff={d:.6f}")

    overall = "PASS" if report["ok"] else "FAIL"
    print(f"\n=== Reconciliation: {overall} ===")
    return 0 if report["ok"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
