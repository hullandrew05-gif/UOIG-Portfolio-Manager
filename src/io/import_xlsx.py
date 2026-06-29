"""Seed the SQLite store from the legacy 'Portfolio Holdings.xlsx' export.

Each fund sheet lays positions out in fixed columns (A=Sector ... N=benchmark
ETF price at entry). A *position* row is identified by having BOTH a Name (col C)
and a Ticker (col D); this cleanly excludes total rows, the embedded
sector-summary block, and blank spacer rows. The parse is verified cell-for-cell
against the workbook by scripts/reconcile.py.

Cash sweeps (BGNXX/BNGXX) carry only a market value in the sheet, so we model
them as a $1.00 NAV money fund: shares == market value, price == 1.0.
"""
from __future__ import annotations

import datetime as dt
import sqlite3
from pathlib import Path

import openpyxl

from src.config import db_path, workbook_path
from src.model import db, schema

# 1-based column indices within each fund sheet
COL = {
    "sector": 1, "cap": 2, "name": 3, "ticker": 4, "shares": 5,
    "price": 6, "mv": 7, "entry_price": 8, "entry_date": 9,
    "passive": 10, "bench_price": 14,
}


def classify(sector: str | None, cap: str | None) -> str:
    s = (sector or "").strip().lower()
    c = (cap or "").strip().lower()
    if s == "cash" or c == "cash":
        return "cash"
    if "index" in s or "index" in c:
        return "etf"
    return "stock"


def _num(v) -> float | None:
    return float(v) if isinstance(v, (int, float)) else None


def _date(v) -> str | None:
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    return None


def parse_fund(ws, fund_name: str, benchmark: str) -> list[dict]:
    """Extract position rows from one fund worksheet."""
    positions: list[dict] = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, COL["name"]).value
        ticker = ws.cell(r, COL["ticker"]).value
        if not name or not ticker:
            continue  # totals, sector summaries, blank rows

        sector = ws.cell(r, COL["sector"]).value
        cap = ws.cell(r, COL["cap"]).value
        sec_type = classify(sector, cap)
        mv = _num(ws.cell(r, COL["mv"]).value)

        if sec_type == "cash":
            shares, price = (mv if mv is not None else 0.0), 1.0
            entry_price, entry_date = 1.0, None
            passive, bench_price, bench = None, None, None
        else:
            shares = _num(ws.cell(r, COL["shares"]).value)
            price = _num(ws.cell(r, COL["price"]).value)
            entry_price = _num(ws.cell(r, COL["entry_price"]).value)
            entry_date = _date(ws.cell(r, COL["entry_date"]).value)
            passive = _num(ws.cell(r, COL["passive"]).value)
            bench_price = _num(ws.cell(r, COL["bench_price"]).value)
            bench = benchmark if sec_type == "stock" else None

        positions.append({
            "fund": fund_name,
            "ticker": str(ticker).strip(),
            "name": str(name).strip(),
            "sector": (sector or "").strip(),
            "cap_class": (cap or "").strip(),
            "sec_type": sec_type,
            "shares": shares,
            "price": price,
            "entry_price": entry_price,
            "entry_date": entry_date,
            "passive_weight": passive,
            "bench_ticker": bench,
            "bench_entry_price": bench_price,
        })
    return positions


def import_workbook(cfg: dict, conn: sqlite3.Connection | None = None) -> dict:
    """(Re)seed the database from the configured workbook. Returns a summary."""
    wb = openpyxl.load_workbook(workbook_path(cfg), data_only=True)

    own_conn = conn is None
    if own_conn:
        conn = schema.get_connection(db_path(cfg))
    schema.create_schema(conn)
    schema.truncate_all(conn)

    import_date = dt.date.today().isoformat()
    cur = conn.cursor()
    seen: dict[str, str] = {}
    summary = {"funds": {}, "warnings": []}

    for fund in cfg["funds"]:
        positions = parse_fund(wb[fund["sheet"]], fund["name"], fund["benchmark"])
        for p in positions:
            tk = p["ticker"]
            if tk not in seen:
                cur.execute(
                    db.q(conn, "INSERT INTO securities (ticker, name, sector, cap_class, sec_type)"
                               " VALUES (?, ?, ?, ?, ?)"),
                    (tk, p["name"], p["sector"], p["cap_class"], p["sec_type"]),
                )
                seen[tk] = p["sector"]
            elif seen[tk] != p["sector"]:
                summary["warnings"].append(
                    f"{tk}: sector differs across funds ('{seen[tk]}' vs '{p['sector']}'); kept first"
                )

            cur.execute(
                db.q(conn, "INSERT INTO holdings (fund, ticker, shares, entry_price, entry_date,"
                           " passive_weight, bench_ticker, bench_entry_price)"
                           " VALUES (?, ?, ?, ?, ?, ?, ?, ?)"),
                (p["fund"], tk, p["shares"], p["entry_price"], p["entry_date"],
                 p["passive_weight"], p["bench_ticker"], p["bench_entry_price"]),
            )

            if p["price"] is not None:
                cur.execute(
                    db.upsert_sql(conn, "prices", ["ticker", "date", "close", "adj_close", "source"], ["ticker", "date"]),
                    (tk, import_date, p["price"], p["price"], "xlsx_snapshot"),
                )

            # snapshot the index ETFs into the benchmarks table too
            if p["sec_type"] == "etf" and p["price"] is not None:
                cur.execute(
                    db.upsert_sql(conn, "benchmarks", ["index_ticker", "date", "close"], ["index_ticker", "date"]),
                    (tk, import_date, p["price"]),
                )

        summary["funds"][fund["name"]] = len(positions)

    for k, v in {
        "import_date": import_date,
        "source_workbook": str(workbook_path(cfg).name),
        "n_securities": str(len(seen)),
    }.items():
        cur.execute(db.upsert_sql(conn, "import_meta", ["key", "value"], ["key"]), (k, v))

    conn.commit()
    if own_conn:
        conn.close()
    summary["n_securities"] = len(seen)
    return summary
